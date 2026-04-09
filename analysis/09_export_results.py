"""
Summary Tables — Final Results Workbook
========================================

Consolidates all key results into a single Excel workbook for presentation.
Each sheet corresponds to one analysis block.

Sheets:
  1. Event Study CAAR    — main CAAR table (US-only primary + full sample)
  2. DID Results         — DID β3 table (US-only primary + full sample)
  3. Placebo DID         — β3 across 6 breakpoints (robustness)
  4. Parallel Trends     — monthly correlation test
  5. Paywall Bounds      — β3 sensitivity to paywall AI rate
  6. Calendar Time       — clustering-corrected FF4 alpha
  7. Repeat Events       — first vs. subsequent event CAAR comparison
  8. Pre-Announcement    — pre-drift test statistics
  9. AI Label Comparison — the three AI definition trials side by side

Run:  python analysis/summary_tables.py
Output: data/results/FINAL_RESULTS.xlsx
"""

import os
import warnings
import numpy as np
import pandas as pd

warnings.filterwarnings('ignore')

BASE     = '/Users/irmina/Documents/Claude/layoff_study'
RES      = os.path.join(BASE, 'data/results')
ROB      = os.path.join(RES, 'robustness')
CT       = os.path.join(RES, 'calendar_time')
IMP      = os.path.join(RES, 'improved')
OUT_PATH = os.path.join(RES, 'FINAL_RESULTS.xlsx')


# ── Helpers ────────────────────────────────────────────────────────────────

def stars(p):
    if pd.isna(p): return ''
    return '***' if p < 0.01 else '**' if p < 0.05 else '*' if p < 0.10 else ''

def fmt_coef(c, p, decimals=3):
    """Return 'X.XXX***' string."""
    if pd.isna(c): return '—'
    return f'{c:.{decimals}f}{stars(p)}'

def fmt_se(se):
    if pd.isna(se): return ''
    return f'({se:.3f})'

def style_sheet(ws, header_color='4472C4'):
    """Apply basic formatting to an openpyxl worksheet."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    thin = Side(style='thin')
    border = Border(bottom=thin)
    fill   = PatternFill('solid', fgColor=header_color)
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
    # Auto-width (approximate)
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Event Study CAAR
# ══════════════════════════════════════════════════════════════════════════════

def sheet_event_study():
    path = os.path.join(RES, 'car_summary.csv')
    if not os.path.exists(path):
        print('  car_summary.csv not found — skipping sheet 1')
        return pd.DataFrame()

    df = pd.read_csv(path)

    # Primary: US only (PRIMARY), FF4 model; secondary: Full sample
    priority = {
        'US only (PRIMARY)': 0,
        'Core tech, US only': 1,
        'Non-tech, US only': 2,
        'Full sample': 3,
        'US only, Post-GenAI (>=2023)': 4,
        'US only, Pre-GenAI (<=2022)': 5,
        'Post-GenAI (>=2023)': 6,
        'Pre-GenAI (<=2022)': 7,
    }
    df['sort_key'] = df['sample'].map(priority).fillna(99)
    df = df[df['model'] == 'FF4'].sort_values(['sort_key', 'window'])

    # Build presentation table
    rows = []
    for _, r in df.iterrows():
        rows.append({
            'Sample':      r['sample'],
            'Model':       r['model'],
            'Window':      r['window'],
            'N':           int(r['N']),
            'CAAR (%)':    f"{r['CAAR_pct']:.3f}{stars(min(p for p in [r['p_patell'],r['p_BMP'],r['p_corrado']] if not pd.isna(p)) if any(not pd.isna(p) for p in [r['p_patell'],r['p_BMP'],r['p_corrado']]) else np.nan)}",
            'Patell t':    f"{r['t_patell']:.3f}" if not pd.isna(r['t_patell']) else '—',
            'p (Patell)':  f"{r['p_patell']:.4f}" if not pd.isna(r['p_patell']) else '—',
            'BMP t':       f"{r['t_BMP']:.3f}"    if not pd.isna(r['t_BMP'])    else '—',
            'p (BMP)':     f"{r['p_BMP']:.4f}"    if not pd.isna(r['p_BMP'])    else '—',
            'Corrado t':   f"{r['t_corrado']:.3f}" if not pd.isna(r['t_corrado']) else '—',
            'p (Corrado)': f"{r['p_corrado']:.4f}" if not pd.isna(r['p_corrado']) else '—',
            'Sig':         r.get('stars', ''),
        })
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — DID Results
# ══════════════════════════════════════════════════════════════════════════════

def sheet_did():
    rows = []
    for fname, sample_label in [
        ('did_results_us_primary.csv', 'US only (PRIMARY)'),
        ('did_results_core_tech.csv',  'Core tech, US only'),
        ('did_results.csv',            'Full sample'),
    ]:
        path = os.path.join(IMP, fname)
        if not os.path.exists(path):
            continue
        df = pd.read_csv(path)
        df['sample_label'] = sample_label
        rows.append(df)

    if not rows:
        print('  DID result files not found')
        return pd.DataFrame()

    df = pd.concat(rows, ignore_index=True)

    # Build clean table focusing on key variables
    key_vars = ['ai_mentioned', 'post_chatgpt', 'ai_x_post', 'const']
    out_rows = []
    for (outcome, spec, sample_lbl), grp in df.groupby(['outcome','spec','sample_label']):
        row = {'Outcome': outcome, 'Specification': spec, 'Sample': sample_lbl}
        for var in key_vars:
            sub = grp[grp['variable'] == var]
            if sub.empty:
                row[var + ' coef'] = '—'
                row[var + ' SE']   = ''
            else:
                r = sub.iloc[0]
                row[var + ' coef'] = fmt_coef(r['coef'], r['pval'])
                row[var + ' SE']   = fmt_se(r['se'])
        sub_n = grp[grp['variable'] == key_vars[0]]
        row['N']  = int(sub_n['N'].iloc[0])  if len(sub_n) else '—'
        row['R²'] = f"{sub_n['R2'].iloc[0]:.3f}" if len(sub_n) else '—'
        out_rows.append(row)

    return pd.DataFrame(out_rows)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Placebo DID
# ══════════════════════════════════════════════════════════════════════════════

def sheet_placebo():
    path = os.path.join(ROB, 'placebo_did_results.csv')
    if not os.path.exists(path):
        return pd.DataFrame()
    df = pd.read_csv(path)
    df_us = df[df['sample'] == 'US only'].copy()
    df_us['Real Breakpoint?'] = df_us['is_real'].map({True: 'YES ★', False: 'No'})
    df_us['β3 (DID)'] = df_us.apply(lambda r: fmt_coef(r['beta3'], r['pval']), axis=1)
    df_us['SE']       = df_us['se'].apply(fmt_se)
    df_us['p-value']  = df_us['pval'].apply(lambda x: f'{x:.4f}' if not pd.isna(x) else '—')
    return df_us[['breakpoint','Real Breakpoint?','β3 (DID)','SE','p-value','N']].rename(
        columns={'breakpoint': 'Breakpoint Date'})


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — Parallel Trends
# ══════════════════════════════════════════════════════════════════════════════

def sheet_parallel_trends():
    path = os.path.join(ROB, 'parallel_trends_monthly.csv')
    if not os.path.exists(path):
        return pd.DataFrame()
    df = pd.read_csv(path)
    df = df.rename(columns={'year_month':'Month','mean_ai':'Mean CAR AI=1 (%)','mean_nai':'Mean CAR AI=0 (%)'})
    df['Mean CAR AI=1 (%)'] = df['Mean CAR AI=1 (%)'].round(3)
    df['Mean CAR AI=0 (%)'] = df['Mean CAR AI=0 (%)'].round(3)
    return df


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5 — Paywall Bounds
# ══════════════════════════════════════════════════════════════════════════════

def sheet_paywall():
    path = os.path.join(ROB, 'paywall_sensitivity.csv')
    if not os.path.exists(path):
        return pd.DataFrame()
    df = pd.read_csv(path)
    df['Assumed Paywall AI Rate (%)'] = (df['assumed_paywall_ai_rate'] * 100).round(1)
    df['Mean β3'] = df['mean_beta3'].round(3)
    df['Std β3']  = df['std_beta3'].round(3)
    df['Mean p']  = df['mean_pval'].round(4)
    df['% Sims Sig (p<5%)'] = df['pct_sig_5pct'].round(1)
    df['Note'] = df['assumed_paywall_ai_rate'].apply(
        lambda x: '← Baseline (no correction)' if x == 0.0
        else ('← Observed readable rate' if abs(x - 0.357) < 0.01 else ''))
    return df[['Assumed Paywall AI Rate (%)','Mean β3','Std β3','Mean p','% Sims Sig (p<5%)','Note']]


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 6 — Calendar Time
# ══════════════════════════════════════════════════════════════════════════════

def sheet_calendar_time():
    path = os.path.join(CT, 'ct_results.csv')
    if not os.path.exists(path):
        return pd.DataFrame()
    df = pd.read_csv(path)
    df['α/month (%)']     = df['alpha_monthly_pct'].round(3)
    df['α annualized (%)']= df['alpha_annual_pct'].round(2)
    df['t (α)']           = df['t_alpha'].round(3)
    df['p (α)']           = df['p_alpha'].round(4)
    df['Sig']             = df['p_alpha'].apply(stars)
    df['Avg Firms/Month'] = df['avg_firms_month'].round(1)
    df['Months']          = df['n_months'].astype(int)
    df['R²']              = df['R2'].round(3)
    df['Note'] = 'Clustering-corrected (Jaffe/Fama method)'
    return df[['label','Months','Avg Firms/Month','α/month (%)','t (α)','p (α)','Sig','α annualized (%)','R²','Note']].rename(
        columns={'label':'Subsample'})


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 7 — Repeat Events
# ══════════════════════════════════════════════════════════════════════════════

def sheet_repeat_events():
    path = os.path.join(ROB, 'repeat_events_summary.csv')
    if not os.path.exists(path):
        return pd.DataFrame()
    df = pd.read_csv(path)
    for col in ['mean_car','se','caar','se']:
        if col in df.columns:
            df[col] = df[col].round(3)
    return df


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 8 — Pre-Announcement Drift
# ══════════════════════════════════════════════════════════════════════════════

def sheet_size_sector():
    path = os.path.join(RES, 'size_sector', 'size_sector_caar.csv')
    if not os.path.exists(path):
        print('    size_sector_caar.csv not found — run size_sector_analysis.py first')
        return pd.DataFrame()
    df = pd.read_csv(path)
    df['CAAR_pct'] = df['CAAR_pct'].round(3)
    df['t']        = df['t'].round(3)
    df['p']        = df['p'].round(4)
    return df.rename(columns={
        'window': 'Window', 'size_group': 'Size Tier (R² proxy)',
        'sector': 'Sector', 'CAAR_pct': 'CAAR (%)', 'sig': 'Sig'
    })


def sheet_pre_announcement():
    path = os.path.join(ROB, 'pre_announcement_stats.csv')
    if not os.path.exists(path):
        return pd.DataFrame()
    df = pd.read_csv(path)
    df['mean_car_pct'] = df['mean_car_pct'].round(3)
    df['se']           = df['se'].round(3)
    df['t']            = df['t'].round(3)
    df['p']            = df['p'].round(4)
    df['Interpretation'] = df['p'].apply(
        lambda p: 'Significant pre-drift (leakage concern)' if not pd.isna(p) and p < 0.05
        else ('Marginal' if not pd.isna(p) and p < 0.10 else 'No significant pre-drift'))
    return df.rename(columns={'window':'Window','mean_car_pct':'Mean CAR (%)','se':'SE',
                               't':'t-stat','p':'p-value','stars':'Sig'})


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 9 — AI Label Comparison
# ══════════════════════════════════════════════════════════════════════════════

def sheet_ai_definitions():
    rows = [
        {'Definition':        'Trial 1: Broad Keyword (Desktop)',
         'Method':            'Case-insensitive substring: AI, efficiency, automation, …',
         'N (ai=1)':          234,
         'Rate (%)':          50.1,
         'Main Issue':        'Substring "ai" matches "said","paid"; "efficiency" in every article',
         'False Positive Est':  '~30–40%',
         'Verdict':           'Rejected — over-labeled'},
        {'Definition':        'Trial 2: Strict Causal Regex',
         'Method':            'Explicit causal phrases only (e.g. "replacing workers due to AI")',
         'N (ai=1)':          12,
         'Rate (%)':          2.6,
         'Main Issue':        'Misses soft language ("transformation","AI strategy")',
         'False Positive Est':  '~0% (high precision, low recall)',
         'Verdict':           'Rejected — severely under-labeled'},
        {'Definition':        'Trial 3a: ai_causal (T3 only)',
         'Method':            'Explicit causal statement — whole-word regex',
         'N (ai=1)':          9,
         'Rate (%)':          1.9,
         'Main Issue':        'Too few events for regression',
         'False Positive Est':  'Very low',
         'Verdict':           'Robustness upper bound only'},
        {'Definition':        'Trial 3b: ai_primary (T2+T3)',
         'Method':            'Specific GenAI tools cited (ChatGPT, LLM, Copilot, …)',
         'N (ai=1)':          22,
         'Rate (%)':          4.7,
         'Main Issue':        'Small N, especially pre-ChatGPT (N=5)',
         'False Positive Est':  'Low',
         'Verdict':           'Robustness check'},
        {'Definition':        'Trial 3c: ai_broad (T1+T2+T3) ★ RECOMMENDED',
         'Method':            'Any specific AI technology term — whole-word only',
         'N (ai=1)':          74,
         'Rate (%)':          15.8,
         'Main Issue':        '42% paywall → systematic undercounting',
         'False Positive Est':  '~10–15% (manageable)',
         'Verdict':           'Primary variable — best balance of precision and power'},
    ]
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# Main — write Excel
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print('=' * 65)
    print('BUILDING FINAL RESULTS WORKBOOK')
    print('=' * 65)

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        has_openpyxl = True
    except ImportError:
        has_openpyxl = False
        print('  openpyxl not installed — will save as CSV bundle instead')

    sheets = {
        '1_EventStudy_CAAR':     sheet_event_study,
        '2_DID_Results':         sheet_did,
        '3_Placebo_DID':         sheet_placebo,
        '4_Parallel_Trends':     sheet_parallel_trends,
        '5_Paywall_Bounds':      sheet_paywall,
        '6_CalendarTime':        sheet_calendar_time,
        '7_RepeatEvents':        sheet_repeat_events,
        '8_PreAnnouncement':     sheet_pre_announcement,
        '9_SizeSector':          sheet_size_sector,
        '10_AI_Definitions':     sheet_ai_definitions,
    }

    dfs = {}
    for name, fn in sheets.items():
        print(f'  Building sheet: {name}...')
        try:
            df = fn()
            dfs[name] = df
            print(f'    → {len(df)} rows')
        except Exception as e:
            print(f'    ⚠ Error: {e}')
            dfs[name] = pd.DataFrame()

    if has_openpyxl:
        with pd.ExcelWriter(OUT_PATH, engine='openpyxl') as writer:
            for name, df in dfs.items():
                if df.empty:
                    # Write placeholder
                    pd.DataFrame([{'Note': 'Data not available — run the relevant analysis script first.'}])\
                      .to_excel(writer, sheet_name=name[:31], index=False)
                else:
                    df.to_excel(writer, sheet_name=name[:31], index=False)

            # Style each sheet
            wb = writer.book
            for ws in wb.worksheets:
                style_sheet(ws)

        print(f'\n  Excel workbook saved → {OUT_PATH}')
    else:
        # Fallback: save each sheet as CSV
        bundle_dir = os.path.join(RES, 'summary_bundle')
        os.makedirs(bundle_dir, exist_ok=True)
        for name, df in dfs.items():
            if not df.empty:
                df.to_csv(os.path.join(bundle_dir, f'{name}.csv'), index=False)
        print(f'\n  CSVs saved → {bundle_dir}')

    # Print a compact text summary for the console
    print('\n' + '=' * 65)
    print('FINAL STUDY SUMMARY')
    print('=' * 65)

    # Sheet 1 key rows
    try:
        caar_df = dfs.get('1_EventStudy_CAAR', pd.DataFrame())
        us_caar = caar_df[caar_df['Sample'].str.contains('US only', na=False)] if not caar_df.empty else pd.DataFrame()
        if not us_caar.empty:
            print('\n  Event Study CAAR (US-only, FF4):')
            for _, r in us_caar.iterrows():
                print(f'    {r["Window"]:<14} CAAR={r["CAAR (%)"]:<12} '
                      f'Patell t={r["Patell t"]:<8} BMP t={r["BMP t"]}')
    except Exception:
        pass

    # Sheet 2 DID β3
    try:
        did_df = dfs.get('2_DID_Results', pd.DataFrame())
        if not did_df.empty:
            print('\n  DID β3 (AI × Post-ChatGPT):')
            for _, r in did_df.iterrows():
                if 'ai_x_post coef' in r.index:
                    print(f'    {r.get("Sample",""):<22} {r.get("Outcome",""):<14} '
                          f'β3={r["ai_x_post coef"]:<12} SE={r.get("ai_x_post SE","")}')
    except Exception:
        pass

    # Placebo
    try:
        pl_df = dfs.get('3_Placebo_DID', pd.DataFrame())
        real_row = pl_df[pl_df['Real Breakpoint?'].str.contains('YES', na=False)] if not pl_df.empty else pd.DataFrame()
        if not real_row.empty:
            r = real_row.iloc[0]
            print(f'\n  Placebo DID — Real breakpoint β3: {r["β3 (DID)"]}  p={r["p-value"]}')
    except Exception:
        pass

    # Parallel trends
    try:
        pt_df = dfs.get('4_Parallel_Trends', pd.DataFrame())
        if not pt_df.empty:
            from scipy.stats import pearsonr
            r_val, p_val = pearsonr(pt_df['Mean CAR AI=1 (%)'], pt_df['Mean CAR AI=0 (%)'])
            verdict = 'PLAUSIBLE' if r_val > 0.4 else 'INCONCLUSIVE'
            print(f'  Parallel Trends — r={r_val:.3f}  p={p_val:.3f}  → {verdict}')
    except Exception:
        pass

    # Calendar time
    try:
        ct_df = dfs.get('6_CalendarTime', pd.DataFrame())
        if not ct_df.empty:
            print('\n  Calendar-Time α (clustering-corrected):')
            for _, r in ct_df.iterrows():
                print(f'    {r["Subsample"]:<35} α={r["α/month (%)"]:.3f}%/mo  '
                      f't={r["t (α)"]:.2f}  {r["Sig"]}')
    except Exception:
        pass

    print(f'\n  Full workbook: {OUT_PATH}')
    print('=' * 65)


if __name__ == '__main__':
    main()
